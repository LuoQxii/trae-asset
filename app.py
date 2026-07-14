"""
个人多账户资产管理系统 - 后端 Flask API v5.0
SQLite 数据库 + 本金/收益双台账体系 + 投资台账 + 公式引擎 + 标的库
"""
import os
import sqlite3
import hashlib
import io
import base64
from datetime import datetime, timedelta
from collections import OrderedDict
from flask import Flask, request, jsonify, send_from_directory, g

app = Flask(__name__)

# 手动 CORS 处理，无需 flask-cors 依赖
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    return response

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, 'asset.db')

transfer_type_labels = {
    'cash_transfer': '现金互转',
    'cash_to_fund': '现金转理财',
    'fund_to_cash': '理财转现金',
    'fund_transfer': '理财互转'
}

# ========== 递增ID生成 ==========
SEQ_ID_LOCK = {}

def generate_seq_id(table_name, db=None):
    """生成5位递增数字ID，从00001开始"""
    if db is None:
        db = get_db()
    # 查询当前表中最大的数字ID
    rows = db.execute(f"SELECT id FROM {table_name} WHERE id GLOB '[0-9]*'").fetchall()
    max_id = 0
    for r in rows:
        try:
            max_id = max(max_id, int(r['id']))
        except (ValueError, TypeError):
            pass
    return f"{max_id + 1:05d}"

# ========== 数据库连接 ==========
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY, username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL, created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS accounts (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            name TEXT NOT NULL, type TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0, created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS principals (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            account_id TEXT NOT NULL, amount REAL NOT NULL,
            source_type TEXT NOT NULL DEFAULT 'other',
            note TEXT DEFAULT '', created_at TEXT NOT NULL,
            record_date TEXT DEFAULT '',
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(account_id) REFERENCES accounts(id)
        );
        CREATE TABLE IF NOT EXISTS returns (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            account_id TEXT NOT NULL, amount REAL NOT NULL,
            return_type TEXT NOT NULL DEFAULT '', note TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            record_date TEXT DEFAULT '',
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(account_id) REFERENCES accounts(id)
        );
        CREATE TABLE IF NOT EXISTS transfers (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            from_account_id TEXT NOT NULL, to_account_id TEXT NOT NULL,
            amount REAL NOT NULL, fee REAL NOT NULL DEFAULT 0,
            transfer_type TEXT NOT NULL DEFAULT 'cash_transfer',
            note TEXT DEFAULT '', created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS snapshots (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            date TEXT NOT NULL, time TEXT NOT NULL,
            total_principal REAL NOT NULL, total_return REAL NOT NULL,
            total_assets REAL NOT NULL, total_debt REAL NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE INDEX IF NOT EXISTS idx_accounts_user ON accounts(user_id);
        CREATE INDEX IF NOT EXISTS idx_principals_user ON principals(user_id);
        CREATE INDEX IF NOT EXISTS idx_principals_account ON principals(account_id);
        CREATE INDEX IF NOT EXISTS idx_returns_user ON returns(user_id);
        CREATE INDEX IF NOT EXISTS idx_returns_account ON returns(account_id);
        CREATE INDEX IF NOT EXISTS idx_transfers_user ON transfers(user_id);
        CREATE INDEX IF NOT EXISTS idx_snapshots_user ON snapshots(user_id);
        CREATE TABLE IF NOT EXISTS dict_items (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            dict_type TEXT NOT NULL,
            label TEXT NOT NULL, value TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0, created_at TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_dict_items_user_type ON dict_items(user_id, dict_type);
        CREATE TABLE IF NOT EXISTS investments (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            account_id TEXT NOT NULL,
            name TEXT NOT NULL,
            investment_type TEXT NOT NULL DEFAULT '',
            direction TEXT NOT NULL DEFAULT '',
            amount REAL NOT NULL DEFAULT 0,
            start_date TEXT DEFAULT '',
            end_date TEXT DEFAULT '',
            status TEXT NOT NULL DEFAULT 'active',
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(account_id) REFERENCES accounts(id)
        );
        CREATE INDEX IF NOT EXISTS idx_investments_user ON investments(user_id);
        CREATE INDEX IF NOT EXISTS idx_investments_account ON investments(account_id);
        CREATE INDEX IF NOT EXISTS idx_investments_type ON investments(investment_type);
        CREATE TABLE IF NOT EXISTS formulas (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            name TEXT NOT NULL,
            formula TEXT NOT NULL DEFAULT '',
            target_field TEXT NOT NULL DEFAULT '',
            description TEXT DEFAULT '',
            is_active INTEGER NOT NULL DEFAULT 1,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE INDEX IF NOT EXISTS idx_formulas_user ON formulas(user_id);
        CREATE TABLE IF NOT EXISTS securities (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
            code TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            security_type TEXT NOT NULL DEFAULT '',
            exchange TEXT NOT NULL DEFAULT '',
            description TEXT DEFAULT '',
            decimals INTEGER NOT NULL DEFAULT 2,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );
        CREATE INDEX IF NOT EXISTS idx_securities_user ON securities(user_id);
        CREATE INDEX IF NOT EXISTS idx_securities_type ON securities(security_type);
    """)

    # ---- 迁移逻辑：为 principals / returns 添加 record_date 列 ----
    cols = [r[1] for r in db.execute("PRAGMA table_info(principals)").fetchall()]
    if 'record_date' not in cols:
        db.execute("ALTER TABLE principals ADD COLUMN record_date TEXT DEFAULT ''")
        db.execute("UPDATE principals SET record_date = substr(created_at, 1, 10) WHERE record_date = '' OR record_date IS NULL")
        db.commit()

    cols = [r[1] for r in db.execute("PRAGMA table_info(returns)").fetchall()]
    if 'record_date' not in cols:
        db.execute("ALTER TABLE returns ADD COLUMN record_date TEXT DEFAULT ''")
        db.execute("UPDATE returns SET record_date = substr(created_at, 1, 10) WHERE record_date = '' OR record_date IS NULL")
        db.commit()

    # ---- 迁移逻辑：为 principals / returns 添加 investment_id 列 ----
    cols = [r[1] for r in db.execute("PRAGMA table_info(principals)").fetchall()]
    if 'investment_id' not in cols:
        db.execute("ALTER TABLE principals ADD COLUMN investment_id TEXT DEFAULT ''")
        db.commit()

    cols = [r[1] for r in db.execute("PRAGMA table_info(returns)").fetchall()]
    if 'investment_id' not in cols:
        db.execute("ALTER TABLE returns ADD COLUMN investment_id TEXT DEFAULT ''")
        db.commit()

    # ---- 迁移逻辑：为 principals 添加 transfer_id 列 ----
    cols = [r[1] for r in db.execute("PRAGMA table_info(principals)").fetchall()]
    if 'transfer_id' not in cols:
        db.execute("ALTER TABLE principals ADD COLUMN transfer_id TEXT DEFAULT ''")
        db.commit()

    # ---- 创建 investment_id 索引 ----
    db.execute("CREATE INDEX IF NOT EXISTS idx_principals_investment ON principals(investment_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_returns_investment ON returns(investment_id)")
    db.commit()

    # ---- 迁移逻辑：为 accounts 添加 open_date 和 remark 列 ----
    cols = [r[1] for r in db.execute("PRAGMA table_info(accounts)").fetchall()]
    if 'open_date' not in cols:
        db.execute("ALTER TABLE accounts ADD COLUMN open_date TEXT DEFAULT ''")
        db.commit()
    if 'remark' not in cols:
        db.execute("ALTER TABLE accounts ADD COLUMN remark TEXT DEFAULT ''")
        db.commit()
    if 'investment_amount' not in cols:
        db.execute("ALTER TABLE accounts ADD COLUMN investment_amount REAL DEFAULT 0")
        db.commit()

    # ---- 迁移逻辑：为 investments 添加 security_id, cost, quantity 列 ----
    cols = [r[1] for r in db.execute("PRAGMA table_info(investments)").fetchall()]
    if 'security_id' not in cols:
        db.execute("ALTER TABLE investments ADD COLUMN security_id TEXT DEFAULT ''")
        db.commit()
    if 'cost' not in cols:
        db.execute("ALTER TABLE investments ADD COLUMN cost REAL NOT NULL DEFAULT 0")
        db.commit()
    if 'quantity' not in cols:
        db.execute("ALTER TABLE investments ADD COLUMN quantity REAL NOT NULL DEFAULT 0")
        db.commit()

    # ---- 迁移逻辑：为 securities 添加 rate 列 ----
    cols = [r[1] for r in db.execute("PRAGMA table_info(securities)").fetchall()]
    if 'rate' not in cols:
        db.execute("ALTER TABLE securities ADD COLUMN rate REAL NOT NULL DEFAULT 0")
        db.commit()

    # ---- 迁移逻辑：为 investments 添加 remaining_quantity 列 ----
    cols = [r[1] for r in db.execute("PRAGMA table_info(investments)").fetchall()]
    if 'remaining_quantity' not in cols:
        db.execute("ALTER TABLE investments ADD COLUMN remaining_quantity REAL NOT NULL DEFAULT 0")
        db.commit()
        # 将已有数据的remaining_quantity初始化为quantity
        db.execute("UPDATE investments SET remaining_quantity = quantity WHERE remaining_quantity = 0")
        db.commit()

    # ---- 插入默认字典数据（按类型检查，避免重复） ----
    now_iso = datetime.now().isoformat()
    system_user = '__system__'
    default_dicts = [
        # account_type
        ('account_type', 'bank', '银行存款', 1),
        ('account_type', 'fund', '基金理财', 2),
        ('account_type', 'stock', '股票账户', 3),
        ('account_type', 'cash', '现金零钱', 4),
        ('account_type', 'debt', '信贷负债', 5),
        # principal_income
        ('principal_income', 'initial', '初始存量', 1),
        ('principal_income', 'salary', '工资收入', 2),
        ('principal_income', 'side_income', '兼职收入', 3),
        ('principal_income', 'red_packet', '红包', 4),
        ('principal_income', 'gift', '馈赠', 5),
        ('principal_income', 'transfer', '账户划转', 6),
        ('principal_income', 'investment_income', '投资回收', 7),
        ('principal_income', 'other', '其他', 8),
        # principal_spend
        ('principal_spend', 'spend', '消费支出', 1),
        ('principal_spend', 'transfer_out', '转账给他人', 2),
        ('principal_spend', 'transfer', '账户划转', 3),
        ('principal_spend', 'repay', '还信用卡', 4),
        ('principal_spend', 'investment_spend', '投资支出', 5),
        ('principal_spend', 'other', '其他', 6),
        # return_type
        ('return_type', 'fund_dividend', '基金分红', 1),
        ('return_type', 'interest', '理财利息', 2),
        ('return_type', 'stock_gain', '股票浮盈', 3),
        ('return_type', 'deposit_interest', '存款利息', 4),
        ('return_type', 'fund_loss', '理财亏损', 5),
        ('return_type', 'stock_loss', '股票浮亏', 6),
        # transfer_type
        ('transfer_type', 'cash_transfer', '现金互转', 1),
        ('transfer_type', 'cash_to_fund', '现金转理财', 2),
        ('transfer_type', 'fund_to_cash', '理财转现金', 3),
        ('transfer_type', 'fund_transfer', '理财互转', 4),
        # investment_type
        ('investment_type', 'stock', '股票', 1),
        ('investment_type', 'fund', '基金', 2),
        ('investment_type', 'bond', '债券', 3),
        ('investment_type', 'wealth', '理财', 4),
        ('investment_type', 'futures', '期货', 5),
        ('investment_type', 'forex', '外汇', 6),
        ('investment_type', 'deposit', '存款', 7),
        ('investment_type', 'other', '其他', 8),
        # investment_direction
        ('investment_direction', 'buy', '买入', 1),
        ('investment_direction', 'fixed', '定投', 2),
        ('investment_direction', 'other', '其他', 3),
        # investment_status
        ('investment_status', 'active', '进行中', 1),
        ('investment_status', 'partial', '部分减仓', 2),
        ('investment_status', 'closed', '已清仓', 3),
        # security_type
        ('security_type', 'stock', '股票', 1),
        ('security_type', 'fund', '基金', 2),
        ('security_type', 'wealth', '理财', 3),
        ('security_type', 'bond', '债券', 4),
        ('security_type', 'deposit', '存款', 5),
    ]
    for dict_type, value, label, sort_order in default_dicts:
        existing = db.execute(
            "SELECT COUNT(*) FROM dict_items WHERE dict_type=? AND value=? AND user_id='__system__'",
            (dict_type, value)
        ).fetchone()[0]
        if existing == 0:
            try:
                db.execute(
                    "INSERT INTO dict_items (id, user_id, dict_type, label, value, sort_order, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (generate_seq_id('dict_items', db), system_user, dict_type, label, value, sort_order, now_iso)
                )
            except Exception:
                pass
    db.commit()

    # ---- 清理已废弃的方向字典项 ----
    db.execute("DELETE FROM dict_items WHERE dict_type='investment_direction' AND value IN ('sell', 'redeem')")
    db.commit()

    # ---- 插入内置公式默认数据（按 name+user_id='__system__' 检查，避免重复） ----
    built_in_formulas = [
        ('总资产', 'total_principal + total_return', 'total_assets', '总资产 = 总本金 + 累计收益', 1),
        ('持有收益率', 'total_return / holding_principal * 100', 'holding_rate', '持有收益率 = 累计收益 / 持有本金 × 100%', 2),
        ('年化收益率', 'holding_rate * 365 / days_held', 'annual_rate', '年化收益率 = 持有收益率 × 365 / 持有天数', 3),
    ]
    for name, formula, target_field, description, sort_order in built_in_formulas:
        existing = db.execute(
            "SELECT COUNT(*) FROM formulas WHERE name=? AND user_id='__system__'",
            (name,)
        ).fetchone()[0]
        if existing == 0:
            try:
                db.execute(
                    "INSERT INTO formulas (id, user_id, name, formula, target_field, description, is_active, sort_order, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (generate_seq_id('formulas', db), system_user, name, formula, target_field, description, 1, sort_order, now_iso)
                )
            except Exception:
                pass
    db.commit()

    db.commit()
    db.close()

# ========== 工具函数 ==========
def hash_password(pwd):
    return hashlib.sha256(pwd.encode('utf-8')).hexdigest()

def fmt(val):
    return round(float(val), 2)

def row_to_dict(row):
    return dict(row) if row else None

def rows_to_list(rows):
    return [dict(r) for r in rows]

def _account_balance_at(db, account_id, date_str, current_amount):
    """基于当前余额反推指定日期的账户余额"""
    post_p = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM principals WHERE account_id=? AND record_date > ?",
        (account_id, date_str)
    ).fetchone()[0]
    post_r = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM returns WHERE account_id=? AND record_date > ?",
        (account_id, date_str)
    ).fetchone()[0]
    post_t_out = db.execute(
        "SELECT COALESCE(SUM(amount+fee),0) FROM transfers WHERE from_account_id=? AND substr(created_at,1,10) > ?",
        (account_id, date_str)
    ).fetchone()[0]
    post_t_in = db.execute(
        "SELECT COALESCE(SUM(amount),0) FROM transfers WHERE to_account_id=? AND substr(created_at,1,10) > ?",
        (account_id, date_str)
    ).fetchone()[0]
    return current_amount - post_p - post_r - (post_t_in - post_t_out)

# ========== 统计引擎 ==========
def calc_stats(user_id):
    db = get_db()
    accounts = db.execute("SELECT * FROM accounts WHERE user_id=?", (user_id,)).fetchall()
    principals = db.execute("SELECT * FROM principals WHERE user_id=?", (user_id,)).fetchall()
    returns = db.execute("SELECT * FROM returns WHERE user_id=?", (user_id,)).fetchall()
    investments = db.execute("SELECT * FROM investments WHERE user_id=?", (user_id,)).fetchall()

    # 本金余额按账户统计
    principal_by_account = {}
    for p in principals:
        aid = p['account_id']
        principal_by_account[aid] = principal_by_account.get(aid, 0.0) + float(p['amount'])

    total_principal = 0.0
    total_debt = 0.0
    total_investment_amount = 0.0
    structure = {}
    for a in accounts:
        amt = fmt(principal_by_account.get(a['id'], 0.0))
        inv_amt = fmt(float(a['investment_amount'] or 0))
        total_investment_amount += inv_amt
        if a['type'] == 'debt':
            total_debt += abs(amt)
        else:
            total_principal += amt
            structure[a['type']] = fmt(structure.get(a['type'], 0) + amt)

    # 累计收益 = 已实现盈亏（卖出金额 - 对应成本），不计入总资产避免重复
    inv_rows = db.execute("""
        SELECT i.account_id, i.amount as inv_amount, i.quantity as inv_qty, i.remaining_quantity,
               COALESCE(SUM(p.amount), 0) as total_sell_amount
        FROM investments i
        LEFT JOIN principals p ON i.id = p.investment_id AND p.source_type = 'investment_income'
        WHERE i.user_id = ?
        GROUP BY i.id
    """, (user_id,)).fetchall()
    realized_by_account = {}
    total_realized = 0.0
    for row in inv_rows:
        aid = row['account_id']
        inv_amount = float(row['inv_amount'] or 0)
        inv_qty = float(row['inv_qty'] or 0)
        remaining = float(row['remaining_quantity'] or 0)
        sold_qty = inv_qty - remaining
        total_sell_amount = float(row['total_sell_amount'] or 0)
        if inv_qty > 0 and sold_qty > 0:
            sold_cost = inv_amount / inv_qty * sold_qty
        else:
            sold_cost = 0
        realized = total_sell_amount - sold_cost
        realized_by_account[aid] = realized_by_account.get(aid, 0.0) + realized
        total_realized += realized

    total_return = fmt(total_realized)
    # 总资产 = 本金余额 + 投资金额（累计收益已体现在本金余额+投资金额的变动中，不再单独加一遍）
    total_assets = fmt(total_principal + total_investment_amount)

    holding_principal = total_investment_amount

    holding_rate = round((total_return / holding_principal) * 100, 2) if holding_principal > 0 else 0

    # 以第一条正收益记录的时间作为起始（仅用于年化计算参考）
    pos_returns = [r for r in returns if r['amount'] > 0]
    first_return = min((datetime.fromisoformat(r['created_at']) for r in pos_returns), default=None)
    if first_return and holding_principal > 0:
        days = max((datetime.now() - first_return).days, 1)
        annual_rate = round(holding_rate * (365 / days), 2)
    else:
        annual_rate = 0

    debt_ratio = round((total_debt / total_assets) * 100, 1) if total_assets > 0 else 0

    principal_by_source = {}
    for p in principals:
        key = p['source_type'] if p['source_type'] else 'other'
        principal_by_source[key] = fmt(principal_by_source.get(key, 0) + p['amount'])

    acc_map = {a['id']: a['name'] for a in accounts}

    # 投资相关统计
    investment_count = len(investments)
    active_investments = [i for i in investments if i['status'] == 'active']
    investment_by_type = {}
    for i in investments:
        investment_by_type[i['investment_type']] = investment_by_type.get(i['investment_type'], 0) + 1

    # ========== 动态公式计算 ==========
    # 读取用户自定义公式 + 系统内置公式，按目标字段覆盖硬编码值
    try:
        formulas = db.execute(
            "SELECT * FROM formulas WHERE (user_id=? OR user_id='__system__') AND is_active=1 ORDER BY sort_order ASC",
            (user_id,)
        ).fetchall()
        # 构建公式变量上下文
        formula_ctx = {
            'total_principal': total_principal,
            'total_return': total_return,
            'total_assets': total_assets,
            'total_debt': total_debt,
            'holding_principal': holding_principal,
            'holding_rate': holding_rate,
            'annual_rate': annual_rate,
            'debt_ratio': debt_ratio,
            'account_count': len(accounts),
            'principal_count': len(principals),
            'return_count': len(returns),
            'investment_count': investment_count,
            'active_investment_count': len(active_investments),
            'days_held': max((datetime.now() - first_return).days, 1) if first_return else 1,
        }
        for f in formulas:
            expr = (f['formula'] or '').strip()
            target = (f['target_field'] or '').strip()
            if not expr or not target:
                continue
            try:
                # 安全eval：只允许数学运算和已知变量
                result = eval(expr, {"__builtins__": {}}, formula_ctx)
                if target in formula_ctx:
                    formula_ctx[target] = result
                # 同步更新主统计变量
                if target == 'total_assets':
                    total_assets = fmt(result)
                elif target == 'holding_rate':
                    holding_rate = round(result, 2)
                elif target == 'annual_rate':
                    annual_rate = round(result, 2)
                elif target == 'debt_ratio':
                    debt_ratio = round(result, 1)
            except Exception:
                pass
    except Exception:
        pass

    return {
        'total_principal': total_principal,
        'total_return': total_return,
        'total_assets': total_assets,
        'total_debt': fmt(total_debt),
        'holding_principal': fmt(holding_principal),
        'holding_rate': holding_rate,
        'annual_rate': annual_rate,
        'debt_ratio': debt_ratio,
        'account_count': len(accounts),
        'principal_count': len(principals),
        'return_count': len(returns),
        'investment_count': investment_count,
        'active_investment_count': len(active_investments),
        'structure': structure,
        'principal_by_source': principal_by_source,
        'return_by_account': {acc_map.get(k, '未知'): v for k, v in realized_by_account.items()},
        'first_return_date': first_return.isoformat() if first_return else None,
        'investment_by_type': investment_by_type
    }

# ========== 资产快照 ==========
def record_snapshot(user_id):
    stats = calc_stats(user_id)
    db = get_db()
    db.execute("""
        INSERT INTO snapshots (id, user_id, date, time, total_principal, total_return, total_assets, total_debt, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        generate_seq_id('snapshots'), user_id,
        datetime.now().strftime('%Y-%m-%d'), datetime.now().strftime('%H:%M:%S'),
        stats['total_principal'], stats['total_return'], stats['total_assets'],
        stats['total_debt'], datetime.now().isoformat()
    ))
    db.commit()

# ========== 用户 API ==========
@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'}), 400
    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone()
    if existing:
        return jsonify({'success': False, 'message': '用户名已存在'}), 409
    uid = generate_seq_id('users')
    db.execute("INSERT INTO users (id, username, password, created_at) VALUES (?, ?, ?, ?)",
               (uid, username, hash_password(password), datetime.now().isoformat()))
    db.commit()
    return jsonify({'success': True, 'message': '注册成功', 'user': {'id': uid, 'username': username}})

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    db = get_db()
    existing = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not existing:
        return jsonify({'success': False, 'message': '该用户未注册'}), 401
    if existing['password'] != hash_password(password):
        return jsonify({'success': False, 'message': '密码错误，请重新输入'}), 401
    return jsonify({'success': True, 'message': '登录成功',
                    'user': {'id': existing['id'], 'username': existing['username']}})

# ========== 账户 API ==========
@app.route('/api/accounts', methods=['GET'])
def get_accounts():
    user_id = request.args.get('user_id', '')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    db = get_db()
    accounts = db.execute("SELECT * FROM accounts WHERE user_id=? ORDER BY created_at", (user_id,)).fetchall()
    result = []
    for acc in accounts:
        acc_dict = dict(acc)
        # 本金余额 = 本金流水合计
        principal_sum = db.execute("SELECT COALESCE(SUM(amount), 0) FROM principals WHERE account_id=? AND user_id=?", (acc['id'], user_id)).fetchone()[0]
        acc_dict['principal_balance'] = fmt(float(principal_sum))
        # 投资金额
        acc_dict['investment_amount'] = fmt(float(acc_dict.get('investment_amount', 0) or 0))
        # 累计收益 = 已实现盈亏（卖出金额 - 对应成本）
        inv_rows = db.execute("""
            SELECT i.amount as inv_amount, i.quantity as inv_qty, i.remaining_quantity,
                   COALESCE(SUM(p.amount), 0) as total_sell_amount
            FROM investments i
            LEFT JOIN principals p ON i.id = p.investment_id AND p.source_type = 'investment_income'
            WHERE i.account_id = ? AND i.user_id = ?
            GROUP BY i.id
        """, (acc['id'], user_id)).fetchall()
        total_realized = 0
        for row in inv_rows:
            inv_amount = float(row['inv_amount'] or 0)
            inv_qty = float(row['inv_qty'] or 0)
            remaining = float(row['remaining_quantity'] or 0)
            sold_qty = inv_qty - remaining
            total_sell_amount = float(row['total_sell_amount'] or 0)
            if inv_qty > 0 and sold_qty > 0:
                sold_cost = inv_amount / inv_qty * sold_qty
            else:
                sold_cost = 0
            total_realized += total_sell_amount - sold_cost
        acc_dict['total_return'] = fmt(total_realized)
        # 总资产 = 本金余额 + 投资金额 + 累计收益
        # 总资产 = 本金余额 + 投资金额（累计收益已体现在本金与投资金额的变动中，不重复加）
        acc_dict['net_value'] = fmt(float(principal_sum) + float(acc_dict['investment_amount']))
        # 兼容旧逻辑：amount 显示为本金余额
        acc_dict['amount'] = acc_dict['principal_balance']
        result.append(acc_dict)
    return jsonify({'success': True, 'accounts': result})

@app.route('/api/accounts', methods=['POST'])
def add_account():
    data = request.get_json()
    user_id = data.get('user_id', '')
    name = data.get('name', '').strip()
    acc_type = data.get('type', 'bank')
    amount = fmt(data.get('amount', 0))
    open_date = data.get('open_date', '').strip()
    remark = data.get('remark', '').strip()
    if not user_id or not name:
        return jsonify({'success': False, 'message': '缺少必填字段'}), 400
    if acc_type == 'debt':
        amount = -abs(amount)
    acc_id = generate_seq_id('accounts')
    now = datetime.now().isoformat()
    db = get_db()
    db.execute("INSERT INTO accounts (id, user_id, name, type, amount, open_date, remark, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
               (acc_id, user_id, name, acc_type, amount, open_date, remark, now))
    db.commit()
    record_snapshot(user_id)
    return jsonify({'success': True, 'message': '账户添加成功',
                    'account': {'id': acc_id, 'user_id': user_id, 'name': name, 'type': acc_type, 'amount': amount, 'open_date': open_date, 'remark': remark, 'created_at': now}})

@app.route('/api/accounts/<acc_id>', methods=['PUT'])
def update_account(acc_id):
    data = request.get_json()
    db = get_db()
    acc = db.execute("SELECT * FROM accounts WHERE id=?", (acc_id,)).fetchone()
    if not acc:
        return jsonify({'success': False, 'message': '账户不存在'}), 404
    updates = []
    params = []
    if 'name' in data:
        updates.append("name=?")
        params.append(data['name'])
    if 'type' in data:
        updates.append("type=?")
        params.append(data['type'])
    if 'amount' in data:
        amount = fmt(data['amount'])
        if data.get('type', acc['type']) == 'debt':
            amount = -abs(amount)
        updates.append("amount=?")
        params.append(amount)
    if 'open_date' in data:
        updates.append("open_date=?")
        params.append(data['open_date'])
    if 'remark' in data:
        updates.append("remark=?")
        params.append(data['remark'])
    if updates:
        params.append(acc_id)
        db.execute(f"UPDATE accounts SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
    acc = db.execute("SELECT * FROM accounts WHERE id=?", (acc_id,)).fetchone()
    return jsonify({'success': True, 'message': '更新成功', 'account': row_to_dict(acc)})

@app.route('/api/accounts/<acc_id>', methods=['DELETE'])
def delete_account(acc_id):
    db = get_db()
    acc = db.execute("SELECT * FROM accounts WHERE id=?", (acc_id,)).fetchone()
    user_id = acc['user_id'] if acc else ''
    db.execute("DELETE FROM principals WHERE account_id=?", (acc_id,))
    db.execute("DELETE FROM returns WHERE account_id=?", (acc_id,))
    db.execute("DELETE FROM investments WHERE account_id=?", (acc_id,))
    db.execute("DELETE FROM accounts WHERE id=?", (acc_id,))
    db.commit()
    if user_id:
        record_snapshot(user_id)
    return jsonify({'success': True, 'message': '删除成功，关联本金、收益和投资记录已同步清理'})

# ========== 本金台账 API ==========
@app.route('/api/principals', methods=['GET'])
def get_principals():
    user_id = request.args.get('user_id', '')
    account_id = request.args.get('account_id', '')
    source_type = request.args.get('source_type', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400

    db = get_db()
    query = """
        SELECT p.*, a.name as account_name, i.name as investment_name
        FROM principals p
        LEFT JOIN accounts a ON p.account_id=a.id
        LEFT JOIN investments i ON p.investment_id=i.id
        WHERE p.user_id=?
    """
    params = [user_id]
    if account_id:
        query += " AND p.account_id=?"
        params.append(account_id)
    if source_type:
        query += " AND p.source_type=?"
        params.append(source_type)
    if start_date:
        query += " AND p.record_date>=?"
        params.append(start_date)
    if end_date:
        query += " AND p.record_date<=?"
        params.append(end_date)
    query += " ORDER BY p.record_date DESC, p.created_at DESC"
    principals = db.execute(query, params).fetchall()
    return jsonify({'success': True, 'principals': rows_to_list(principals)})

@app.route('/api/principals', methods=['POST'])
def add_principal():
    data = request.get_json()
    user_id = data.get('user_id', '')
    account_id = data.get('account_id', '')
    amount = fmt(data.get('amount', 0))
    source_type = data.get('source_type', 'other')
    note = data.get('note', '').strip()
    investment_id = (data.get('investment_id') or '').strip()
    if not user_id or not account_id:
        return jsonify({'success': False, 'message': '缺少必填字段'}), 400
    if amount == 0:
        return jsonify({'success': False, 'message': '金额不能为0'}), 400

    db = get_db()
    # 验证 investment_id 若提供
    if investment_id:
        inv = db.execute("SELECT id FROM investments WHERE id=? AND user_id=?", (investment_id, user_id)).fetchone()
        if not inv:
            return jsonify({'success': False, 'message': '关联投资不存在'}), 404

    acc = db.execute("SELECT * FROM accounts WHERE id=? AND user_id=?", (account_id, user_id)).fetchone()
    if acc and acc['type'] != 'debt':
        new_amount = fmt(acc['amount'] + amount)
        if new_amount < 0:
            return jsonify({'success': False, 'message': '账户余额不足，无法支出'}), 400
        db.execute("UPDATE accounts SET amount=? WHERE id=?", (new_amount, account_id))

    pid = generate_seq_id('principals')
    now = datetime.now().isoformat()
    record_date = data.get('record_date', datetime.now().strftime('%Y-%m-%d'))
    db.execute(
        "INSERT INTO principals (id, user_id, account_id, amount, source_type, note, created_at, record_date, investment_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (pid, user_id, account_id, amount, source_type, note, now, record_date, investment_id)
    )
    db.commit()
    record_snapshot(user_id)
    return jsonify({'success': True, 'message': '本金录入成功',
                    'principal': {'id': pid, 'user_id': user_id, 'account_id': account_id,
                                  'amount': amount, 'source_type': source_type, 'note': note,
                                  'created_at': now, 'record_date': record_date, 'investment_id': investment_id}})

@app.route('/api/principals/<p_id>', methods=['DELETE'])
def delete_principal(p_id):
    db = get_db()
    p = db.execute("SELECT * FROM principals WHERE id=?", (p_id,)).fetchone()
    if p:
        acc = db.execute("SELECT * FROM accounts WHERE id=? AND user_id=?", (p['account_id'], p['user_id'])).fetchone()
        if acc and acc['type'] != 'debt':
            db.execute("UPDATE accounts SET amount=? WHERE id=?", (fmt(acc['amount'] - p['amount']), p['account_id']))
        db.execute("DELETE FROM principals WHERE id=?", (p_id,))
        db.commit()
        record_snapshot(p['user_id'])
    return jsonify({'success': True, 'message': '删除成功'})

# ========== 收益台账 API ==========
@app.route('/api/returns', methods=['GET'])
def get_returns():
    user_id = request.args.get('user_id', '')
    account_id = request.args.get('account_id', '')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    db = get_db()
    query = """
        SELECT r.*, a.name as account_name, i.name as investment_name
        FROM returns r
        LEFT JOIN accounts a ON r.account_id=a.id
        LEFT JOIN investments i ON r.investment_id=i.id
        WHERE r.user_id=?
    """
    params = [user_id]
    if account_id:
        query += " AND r.account_id=?"
        params.append(account_id)
    query += " ORDER BY r.record_date DESC, r.created_at DESC"
    returns = db.execute(query, params).fetchall()
    return jsonify({'success': True, 'returns': rows_to_list(returns)})

@app.route('/api/returns', methods=['POST'])
def add_return():
    data = request.get_json()
    user_id = data.get('user_id', '')
    account_id = data.get('account_id', '')
    amount = fmt(data.get('amount', 0))
    return_type = data.get('return_type', '')
    note = data.get('note', '').strip()
    investment_id = data.get('investment_id', '').strip()
    if not user_id or not account_id:
        return jsonify({'success': False, 'message': '缺少必填字段'}), 400
    if amount == 0:
        return jsonify({'success': False, 'message': '收益金额不能为0'}), 400

    db = get_db()
    if investment_id:
        inv = db.execute("SELECT id FROM investments WHERE id=? AND user_id=?", (investment_id, user_id)).fetchone()
        if not inv:
            return jsonify({'success': False, 'message': '关联投资不存在'}), 404

    rid = generate_seq_id('returns')
    now = datetime.now().isoformat()
    record_date = data.get('record_date', datetime.now().strftime('%Y-%m-%d'))
    db.execute(
        "INSERT INTO returns (id, user_id, account_id, amount, return_type, note, created_at, record_date, investment_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (rid, user_id, account_id, amount, return_type, note, now, record_date, investment_id)
    )
    db.commit()
    record_snapshot(user_id)
    return jsonify({'success': True, 'message': '收益录入成功',
                    'return': {'id': rid, 'user_id': user_id, 'account_id': account_id,
                               'amount': amount, 'return_type': return_type, 'note': note,
                               'created_at': now, 'record_date': record_date, 'investment_id': investment_id}})

@app.route('/api/returns/<ret_id>', methods=['DELETE'])
def delete_return(ret_id):
    db = get_db()
    ret = db.execute("SELECT * FROM returns WHERE id=?", (ret_id,)).fetchone()
    if ret:
        db.execute("DELETE FROM returns WHERE id=?", (ret_id,))
        db.commit()
        record_snapshot(ret['user_id'])
    return jsonify({'success': True, 'message': '删除成功'})

# ========== 资产划转 API ==========
@app.route('/api/transfers', methods=['GET'])
def get_transfers():
    user_id = request.args.get('user_id', '')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    db = get_db()
    transfers = db.execute("""
        SELECT t.*, fa.name as from_account_name, ta.name as to_account_name
        FROM transfers t
        LEFT JOIN accounts fa ON t.from_account_id=fa.id
        LEFT JOIN accounts ta ON t.to_account_id=ta.id
        WHERE t.user_id=?
        ORDER BY t.created_at DESC
    """, (user_id,)).fetchall()
    result = rows_to_list(transfers)
    for t in result:
        t['transfer_type_label'] = transfer_type_labels.get(t['transfer_type'], t['transfer_type'])
    return jsonify({'success': True, 'transfers': result})

@app.route('/api/transfers', methods=['POST'])
def add_transfer():
    data = request.get_json()
    user_id = data.get('user_id', '')
    from_account_id = data.get('from_account_id', '')
    to_account_id = data.get('to_account_id', '')
    amount = fmt(data.get('amount', 0))
    fee = fmt(data.get('fee', 0))
    transfer_type = data.get('transfer_type', 'cash_transfer')
    note = data.get('note', '').strip()
    record_date = (data.get('record_date') or '').strip() or datetime.now().strftime('%Y-%m-%d')

    if not user_id or not from_account_id or not to_account_id:
        return jsonify({'success': False, 'message': '缺少必填字段'}), 400
    if from_account_id == to_account_id:
        return jsonify({'success': False, 'message': '转出和转入账户不能相同'}), 400
    if amount <= 0:
        return jsonify({'success': False, 'message': '划转金额必须大于0'}), 400
    if fee < 0:
        return jsonify({'success': False, 'message': '手续费不能为负数'}), 400

    db = get_db()
    from_acc = db.execute("SELECT * FROM accounts WHERE id=? AND user_id=?", (from_account_id, user_id)).fetchone()
    to_acc = db.execute("SELECT * FROM accounts WHERE id=? AND user_id=?", (to_account_id, user_id)).fetchone()
    if not from_acc or not to_acc:
        return jsonify({'success': False, 'message': '账户不存在'}), 404

    # 基于本金流水检查余额
    from_balance = db.execute("SELECT COALESCE(SUM(amount), 0) FROM principals WHERE account_id=? AND user_id=?", (from_account_id, user_id)).fetchone()[0]
    total_out = amount + fee
    if float(from_balance) < total_out:
        return jsonify({'success': False, 'message': f'转出账户余额不足（需 ¥{total_out}，含手续费 ¥{fee}）'}), 400

    tid = generate_seq_id('transfers')
    now = datetime.now().isoformat()
    db.execute("INSERT INTO transfers (id, user_id, from_account_id, to_account_id, amount, fee, transfer_type, note, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
               (tid, user_id, from_account_id, to_account_id, amount, fee, transfer_type, note, now))

    # 记录本金流水：转出（负）
    pid_out = generate_seq_id('principals')
    db.execute(
        "INSERT INTO principals (id, user_id, account_id, amount, source_type, note, created_at, record_date, transfer_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (pid_out, user_id, from_account_id, -abs(total_out), 'transfer', '转出至：' + (to_acc['name'] or '') + (' ' + note if note else ''), now, record_date, tid)
    )
    # 记录本金流水：转入（正）
    pid_in = generate_seq_id('principals')
    db.execute(
        "INSERT INTO principals (id, user_id, account_id, amount, source_type, note, created_at, record_date, transfer_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (pid_in, user_id, to_account_id, abs(amount), 'transfer', '从' + (from_acc['name'] or '') + '转入' + (' ' + note if note else ''), now, record_date, tid)
    )

    db.commit()
    record_snapshot(user_id)
    return jsonify({'success': True, 'message': '划转成功',
                    'transfer': {'id': tid, 'user_id': user_id, 'from_account_id': from_account_id,
                                 'to_account_id': to_account_id, 'amount': amount, 'fee': fee,
                                 'transfer_type': transfer_type, 'note': note, 'created_at': now}})

@app.route('/api/transfers/<t_id>', methods=['DELETE'])
def delete_transfer(t_id):
    db = get_db()
    t = db.execute("SELECT * FROM transfers WHERE id=?", (t_id,)).fetchone()
    if not t:
        return jsonify({'success': False, 'message': '划转记录不存在'}), 404

    # 删除关联的本金流水
    db.execute("DELETE FROM principals WHERE transfer_id=?", (t_id,))
    db.execute("DELETE FROM transfers WHERE id=?", (t_id,))
    db.commit()
    record_snapshot(t['user_id'])
    return jsonify({'success': True, 'message': '划转记录已删除，关联本金流水已清理'})

# ========== 快照 API ==========
@app.route('/api/snapshots', methods=['GET'])
def get_snapshots():
    user_id = request.args.get('user_id', '')
    period = request.args.get('period', 'day')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    db = get_db()
    snapshots = db.execute("SELECT * FROM snapshots WHERE user_id=? ORDER BY created_at", (user_id,)).fetchall()
    result = rows_to_list(snapshots)

    if period == 'day':
        aggregated = result
    else:
        grouped = OrderedDict()
        for s in result:
            dt = datetime.fromisoformat(s['created_at'])
            if period == 'week':
                key = dt.strftime('%Y-W%W')
            elif period == 'month':
                key = dt.strftime('%Y-%m')
            elif period == 'year':
                key = dt.strftime('%Y')
            else:
                key = dt.strftime('%Y-%m-%d')
            grouped[key] = s
        aggregated = list(grouped.values())
    return jsonify({'success': True, 'snapshots': aggregated})

# ========== 统计 API ==========
@app.route('/api/stats', methods=['GET'])
def get_stats():
    user_id = request.args.get('user_id', '')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    stats = calc_stats(user_id)
    return jsonify({'success': True, 'stats': stats})

# ========== 数据字典 API ==========
@app.route('/api/dicts', methods=['GET'])
def get_dicts():
    user_id = request.args.get('user_id', '')
    dict_type = request.args.get('dict_type', '')
    db = get_db()
    query = "SELECT * FROM dict_items WHERE 1=1"
    params = []
    if user_id:
        query += " AND (user_id=? OR user_id='__system__')"
        params.append(user_id)
    if dict_type:
        query += " AND dict_type=?"
        params.append(dict_type)
    query += " ORDER BY sort_order ASC, created_at ASC"
    items = db.execute(query, params).fetchall()
    return jsonify({'success': True, 'items': rows_to_list(items)})

@app.route('/api/dicts', methods=['POST'])
def add_dict():
    data = request.get_json()
    user_id = data.get('user_id', '')
    dict_type = data.get('dict_type', '')
    label = data.get('label', '').strip()
    value = data.get('value', '').strip()
    if not user_id or not dict_type or not label or not value:
        return jsonify({'success': False, 'message': '缺少必填字段'}), 400
    did = generate_seq_id('dict_items')
    now = datetime.now().isoformat()
    db = get_db()
    db.execute("INSERT INTO dict_items (id, user_id, dict_type, label, value, sort_order, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
               (did, user_id, dict_type, label, value, 0, now))
    db.commit()
    return jsonify({'success': True, 'message': '字典项添加成功',
                    'item': {'id': did, 'user_id': user_id, 'dict_type': dict_type, 'label': label, 'value': value, 'sort_order': 0, 'created_at': now}})

@app.route('/api/dicts/<did>', methods=['PUT'])
def update_dict(did):
    data = request.get_json()
    db = get_db()
    d = db.execute("SELECT * FROM dict_items WHERE id=?", (did,)).fetchone()
    if not d:
        return jsonify({'success': False, 'message': '字典项不存在'}), 404
    updates = []
    params = []
    if 'label' in data:
        updates.append("label=?")
        params.append(data['label'].strip())
    if 'value' in data:
        updates.append("value=?")
        params.append(data['value'].strip())
    if 'sort_order' in data:
        updates.append("sort_order=?")
        params.append(int(data['sort_order']))
    if updates:
        params.append(did)
        db.execute(f"UPDATE dict_items SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
    d = db.execute("SELECT * FROM dict_items WHERE id=?", (did,)).fetchone()
    return jsonify({'success': True, 'message': '更新成功', 'dict': row_to_dict(d)})

@app.route('/api/dicts/<did>', methods=['DELETE'])
def delete_dict(did):
    db = get_db()
    d = db.execute("SELECT * FROM dict_items WHERE id=?", (did,)).fetchone()
    if not d:
        return jsonify({'success': False, 'message': '字典项不存在'}), 404
    db.execute("DELETE FROM dict_items WHERE id=?", (did,))
    db.commit()
    return jsonify({'success': True, 'message': '删除成功'})

# ========== 投资 API ==========
@app.route('/api/investments', methods=['GET'])
def get_investments():
    user_id = request.args.get('user_id', '')
    account_id = request.args.get('account_id', '')
    name = request.args.get('name', '').strip()
    investment_type = request.args.get('investment_type', '').strip()
    direction = request.args.get('direction', '').strip()
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400

    db = get_db()
    query = """
        SELECT i.*, a.name as account_name,
               s.code as security_code, s.name as security_name, s.security_type as security_type_name
        FROM investments i
        LEFT JOIN accounts a ON i.account_id=a.id
        LEFT JOIN securities s ON i.security_id=s.id
        WHERE i.user_id=?
    """
    params = [user_id]
    if account_id:
        query += " AND i.account_id=?"
        params.append(account_id)
    if name:
        query += " AND i.name LIKE ?"
        params.append(f'%{name}%')
    if investment_type:
        query += " AND i.investment_type=?"
        params.append(investment_type)
    if direction:
        query += " AND i.direction=?"
        params.append(direction)
    if start_date:
        query += " AND i.start_date>=?"
        params.append(start_date)
    if end_date:
        query += " AND i.end_date<=?"
        params.append(end_date)
    if status:
        query += " AND i.status=?"
        params.append(status)
    if search:
        query += " AND (i.name LIKE ? OR i.investment_type LIKE ? OR i.direction LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%', f'%{search}%'])
    query += " ORDER BY i.created_at DESC"
    investments = db.execute(query, params).fetchall()
    result = rows_to_list(investments)
    for inv in result:
        iid = inv['id']
        inv_status = inv.get('status', '')
        inv_amount = float(inv.get('amount', 0) or 0)
        inv_qty = float(inv.get('quantity', 0) or 0)
        remaining = float(inv.get('remaining_quantity', 0) or 0)
        if inv_status == 'closed':
            # 清仓状态：收益 = 已实现盈亏（卖出金额 - 成本）
            total_sell = db.execute(
                "SELECT COALESCE(SUM(amount), 0) FROM principals WHERE investment_id=? AND source_type='investment_income'",
                (iid,)
            ).fetchone()[0]
            sold_qty = inv_qty - remaining
            if inv_qty > 0 and sold_qty > 0:
                sold_cost = inv_amount / inv_qty * sold_qty
            else:
                sold_cost = 0
            inv['total_return'] = fmt(float(total_sell) - sold_cost)
        else:
            # 进行中/部分卖出：收益 = 浮盈浮亏
            total_return = db.execute(
                "SELECT COALESCE(SUM(amount), 0) FROM returns WHERE investment_id=?",
                (iid,)
            ).fetchone()[0]
            inv['total_return'] = fmt(float(total_return))
    return jsonify({'success': True, 'investments': result})

@app.route('/api/investments', methods=['POST'])
def add_investment():
    data = request.get_json()
    user_id = data.get('user_id', '')
    account_id = data.get('account_id', '')
    name = data.get('name', '').strip()
    investment_type = data.get('investment_type', '').strip()
    direction = data.get('direction', '').strip()
    if not direction or direction == 'long':
        direction = 'buy'
    amount = fmt(data.get('amount', 0))
    if direction in ('buy', 'fixed'):
        amount = abs(amount)
    elif direction in ('sell', 'redeem'):
        amount = -abs(amount)
    start_date = (data.get('start_date') or '').strip()
    end_date = (data.get('end_date') or '').strip()
    status = data.get('status', 'active') or 'active'
    security_id = (data.get('security_id') or '').strip()
    cost = fmt(data.get('cost', 0))
    quantity = fmt(data.get('quantity', 0))
    remaining_quantity = fmt(data.get('remaining_quantity', quantity))
    if not user_id or not account_id or not name:
        return jsonify({'success': False, 'message': '缺少必填字段'}), 400
    db = get_db()
    acc = db.execute("SELECT id FROM accounts WHERE id=? AND user_id=?", (account_id, user_id)).fetchone()
    if not acc:
        return jsonify({'success': False, 'message': '账户不存在'}), 404
    iid = generate_seq_id('investments')
    now = datetime.now().isoformat()
    db.execute(
        "INSERT INTO investments (id, user_id, account_id, name, investment_type, direction, amount, start_date, end_date, status, security_id, cost, quantity, remaining_quantity, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (iid, user_id, account_id, name, investment_type, direction, amount, start_date, end_date, status, security_id, cost, quantity, remaining_quantity, now)
    )
    # 扣除本金余额，记录投资支出流水
    pid = generate_seq_id('principals')
    db.execute(
        "INSERT INTO principals (id, user_id, account_id, amount, source_type, note, created_at, record_date, investment_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (pid, user_id, account_id, -abs(amount), 'investment_spend', '投资支出：' + name, now, (start_date or datetime.now().strftime('%Y-%m-%d')), iid)
    )
    # 更新账户投资金额
    db.execute(
        "UPDATE accounts SET investment_amount = COALESCE(investment_amount, 0) + ? WHERE id = ? AND user_id = ?",
        (amount, account_id, user_id)
    )
    db.commit()
    return jsonify({'success': True, 'message': '投资添加成功',
                    'investment': {'id': iid, 'user_id': user_id, 'account_id': account_id,
                                   'name': name, 'investment_type': investment_type, 'direction': direction,
                                   'amount': amount, 'start_date': start_date, 'end_date': end_date,
                                   'status': status, 'security_id': security_id, 'cost': cost,
                                   'quantity': quantity, 'remaining_quantity': remaining_quantity, 'created_at': now}})

@app.route('/api/investments/<iid>', methods=['PUT'])
def update_investment(iid):
    data = request.get_json()
    db = get_db()
    inv = db.execute("SELECT * FROM investments WHERE id=?", (iid,)).fetchone()
    if not inv:
        return jsonify({'success': False, 'message': '投资不存在'}), 404
    if 'direction' in data and not data['direction'].strip():
        return jsonify({'success': False, 'message': 'direction 不能为空'}), 400
    updates = []
    params = []
    fields = ['name', 'investment_type', 'direction', 'amount', 'start_date', 'end_date', 'status', 'account_id', 'security_id', 'cost', 'quantity', 'remaining_quantity']
    for f in fields:
        if f in data:
            updates.append(f"{f}=?")
            if f in ('amount', 'cost', 'quantity', 'remaining_quantity'):
                params.append(fmt(data[f]))
            else:
                params.append(data[f])
    if updates:
        params.append(iid)
        db.execute(f"UPDATE investments SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
    inv = db.execute("SELECT * FROM investments WHERE id=?", (iid,)).fetchone()
    return jsonify({'success': True, 'message': '更新成功', 'investment': row_to_dict(inv)})

@app.route('/api/investments/<iid>', methods=['DELETE'])
def delete_investment(iid):
    db = get_db()
    inv = db.execute("SELECT * FROM investments WHERE id=?", (iid,)).fetchone()
    if not inv:
        return jsonify({'success': False, 'message': '投资不存在'}), 404
    account_id = inv['account_id']
    user_id = inv['user_id']
    inv_amount = float(inv['amount'] or 0)
    inv_quantity = float(inv['quantity'] or 0)
    remaining = float(inv['remaining_quantity'] or 0)
    # 计算剩余成本
    if inv_quantity > 0:
        remaining_cost = inv_amount / inv_quantity * remaining
    else:
        remaining_cost = inv_amount
    # 删除关联的投资支出本金流水
    db.execute("DELETE FROM principals WHERE investment_id=? AND source_type='investment_spend'", (iid,))
    # 删除关联的投资回收本金流水
    db.execute("DELETE FROM principals WHERE investment_id=? AND source_type='investment_income'", (iid,))
    # 回滚账户投资金额（按剩余成本回滚）
    db.execute(
        "UPDATE accounts SET investment_amount = COALESCE(investment_amount, 0) - ? WHERE id = ? AND user_id = ?",
        (remaining_cost, account_id, user_id)
    )
    # 解除关联的 principals 和 returns
    db.execute("UPDATE principals SET investment_id='' WHERE investment_id=?", (iid,))
    db.execute("UPDATE returns SET investment_id='' WHERE investment_id=?", (iid,))
    db.execute("DELETE FROM investments WHERE id=?", (iid,))
    db.commit()
    return jsonify({'success': True, 'message': '删除成功，关联本金和收益记录已解除绑定'})

# ========== 投资专属台账 API ==========
@app.route('/api/investments/<iid>/ledger', methods=['GET'])
def investment_ledger(iid):
    """获取单个投资的专属台账（投资流水 + 收益流水 + 卖出回收流水）"""
    user_id = request.args.get('user_id', '')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    db = get_db()
    inv = db.execute("SELECT * FROM investments WHERE id=? AND user_id=?", (iid, user_id)).fetchone()
    if not inv:
        return jsonify({'success': False, 'message': '投资不存在'}), 404

    inv_status = inv['status'] or ''
    total_cost = float(inv['amount'] or 0)
    inv_qty = float(inv['quantity'] or 0)
    remaining = float(inv['remaining_quantity'] or 0)

    # 已清仓：收益 = 已实现盈亏（卖出总金额 - 成本）
    if inv_status == 'closed':
        total_sell = db.execute(
            "SELECT COALESCE(SUM(amount), 0) FROM principals WHERE investment_id=? AND source_type='investment_income'",
            (iid,)
        ).fetchone()[0]
        total_sell = float(total_sell or 0)
        sold_qty = inv_qty - remaining
        if inv_qty > 0 and sold_qty > 0:
            sold_cost = total_cost / inv_qty * sold_qty
        else:
            sold_cost = total_cost
        total_pnl = total_sell - sold_cost
        pnl_rate = round(total_pnl / total_cost * 100, 2) if total_cost > 0 else 0
    else:
        # 进行中/部分卖出：收益 = 浮盈浮亏
        total_pnl = db.execute(
            "SELECT COALESCE(SUM(amount), 0) FROM returns WHERE investment_id=? AND user_id=?",
            (iid, user_id)
        ).fetchone()[0]
        total_pnl = float(total_pnl or 0)
        pnl_rate = round(total_pnl / total_cost * 100, 2) if total_cost > 0 else 0

    summary = {
        'total_cost': fmt(total_cost),
        'total_pnl': fmt(total_pnl),
        'pnl_rate': pnl_rate
    }

    inv_flows = []
    inv_row = db.execute("""
        SELECT i.*, a.name as account_name,
               s.code as security_code, s.name as security_name, s.security_type as security_type_name
        FROM investments i
        LEFT JOIN accounts a ON i.account_id=a.id
        LEFT JOIN securities s ON i.security_id=s.id
        WHERE i.id=? AND i.user_id=?
    """, (iid, user_id)).fetchone()
    if inv_row:
        inv_flows.append({
            'id': inv_row['id'],
            'flow_type': 'investment',
            'account_name': inv_row['account_name'] or '',
            'investment_name': inv_row['name'] or '',
            'investment_type': inv_row['investment_type'] or '',
            'direction': inv_row['direction'] or '',
            'amount': float(inv_row['amount']),
            'source_type': '',
            'note': '',
            'record_date': inv_row['start_date'] or inv_row['created_at'],
            'security_code': inv_row['security_code'] or '',
            'security_name': inv_row['security_name'] or '',
            'security_type': inv_row['security_type_name'] or '',
        })

    # 收益流水（仅展示，清仓时不计入总盈亏）
    ret_rows = db.execute("""
        SELECT r.*, a.name as account_name
        FROM returns r
        LEFT JOIN accounts a ON r.account_id=a.id
        WHERE r.investment_id=? AND r.user_id=?
        ORDER BY r.record_date DESC, r.created_at DESC
    """, (iid, user_id)).fetchall()
    for r in ret_rows:
        inv_flows.append({
            'id': r['id'],
            'flow_type': 'return',
            'account_name': r['account_name'] or '',
            'investment_name': inv_row['name'] if inv_row else '',
            'investment_type': inv_row['investment_type'] if inv_row else '',
            'direction': '',
            'amount': float(r['amount']),
            'source_type': r['return_type'] or '',
            'note': r['note'] or '',
            'record_date': r['record_date'] or r['created_at'],
            'security_code': inv_row['security_code'] if inv_row else '',
            'security_name': inv_row['security_name'] if inv_row else '',
            'security_type': inv_row['security_type_name'] if inv_row else '',
        })

    # 卖出/清仓回收流水
    rec_rows = db.execute("""
        SELECT p.*, a.name as account_name
        FROM principals p
        LEFT JOIN accounts a ON p.account_id=a.id
        WHERE p.investment_id=? AND p.source_type='investment_income'
        ORDER BY p.record_date DESC, p.created_at DESC
    """, (iid,)).fetchall()
    for r in rec_rows:
        inv_flows.append({
            'id': r['id'],
            'flow_type': 'recovery',
            'account_name': r['account_name'] or '',
            'investment_name': inv_row['name'] if inv_row else '',
            'investment_type': inv_row['investment_type'] if inv_row else '',
            'direction': '',
            'amount': float(r['amount']),
            'source_type': r['source_type'] or '',
            'note': r['note'] or '',
            'record_date': r['record_date'] or r['created_at'],
            'security_code': inv_row['security_code'] if inv_row else '',
            'security_name': inv_row['security_name'] if inv_row else '',
            'security_type': inv_row['security_type_name'] if inv_row else '',
        })

    inv_flows.sort(key=lambda x: x['record_date'], reverse=True)

    return jsonify({
        'success': True,
        'data': {
            'summary': summary,
            'flows': inv_flows
        }
    })

@app.route('/api/investments/<iid>/sell', methods=['POST'])
def sell_investment(iid):
    """部分卖出：扣减持仓剩余数量，记录本金回收流水"""
    data = request.get_json() or {}
    user_id = data.get('user_id', '')
    sell_qty = float(data.get('sell_quantity', 0))
    sell_price = float(data.get('sell_price', 0))
    sell_amount = float(data.get('sell_amount', 0))
    record_date = (data.get('record_date') or '').strip() or datetime.now().strftime('%Y-%m-%d')
    if not user_id or sell_qty <= 0:
        return jsonify({'success': False, 'message': '参数错误'}), 400
    db = get_db()
    inv = db.execute("SELECT * FROM investments WHERE id=? AND user_id=?", (iid, user_id)).fetchone()
    if not inv:
        return jsonify({'success': False, 'message': '投资不存在'}), 404
    remaining = float(inv['remaining_quantity'] or 0)
    if sell_qty > remaining:
        return jsonify({'success': False, 'message': '卖出数量不可大于剩余数量'}), 400
    new_remaining = remaining - sell_qty
    new_status = 'closed' if new_remaining <= 0 else 'partial'
    db.execute(
        "UPDATE investments SET remaining_quantity=?, status=? WHERE id=? AND user_id=?",
        (fmt(new_remaining), new_status, iid, user_id)
    )
    # 记录本金回收流水
    account_id = inv['account_id']
    inv_name = inv['name'] or ''
    inv_amount = float(inv['amount'] or 0)
    inv_quantity = float(inv['quantity'] or 0)
    if inv_quantity > 0:
        sold_cost = inv_amount / inv_quantity * sell_qty
    else:
        sold_cost = sell_amount
    pid = generate_seq_id('principals')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "INSERT INTO principals (id, user_id, account_id, amount, source_type, note, created_at, record_date, investment_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (pid, user_id, account_id, abs(sell_amount), 'investment_income', '投资回收：' + inv_name + ' 部分卖出 ' + str(fmt(sell_qty)) + ' 股', now, record_date, iid)
    )
    # 更新账户投资金额
    db.execute(
        "UPDATE accounts SET investment_amount = COALESCE(investment_amount, 0) - ? WHERE id = ? AND user_id = ?",
        (sold_cost, account_id, user_id)
    )
    db.commit()
    record_snapshot(user_id)
    return jsonify({'success': True, 'message': '卖出成功', 'remaining_quantity': fmt(new_remaining), 'status': new_status})

@app.route('/api/investments/<iid>/close', methods=['POST'])
def close_investment(iid):
    """全额清仓：记录本金回收流水"""
    data = request.get_json() or {}
    user_id = data.get('user_id', '')
    close_price = float(data.get('close_price', 0))
    close_amount = float(data.get('amount', 0))
    record_date = (data.get('record_date') or '').strip() or datetime.now().strftime('%Y-%m-%d')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少user_id'}), 400
    db = get_db()
    inv = db.execute("SELECT * FROM investments WHERE id=? AND user_id=?", (iid, user_id)).fetchone()
    if not inv:
        return jsonify({'success': False, 'message': '投资不存在'}), 404
    db.execute(
        "UPDATE investments SET remaining_quantity=0, status='closed' WHERE id=? AND user_id=?",
        (iid, user_id)
    )
    # 记录本金回收流水
    account_id = inv['account_id']
    inv_name = inv['name'] or ''
    inv_amount = float(inv['amount'] or 0)
    inv_quantity = float(inv['quantity'] or 0)
    remaining = float(inv['remaining_quantity'] or 0)
    if inv_quantity > 0:
        remaining_cost = inv_amount / inv_quantity * remaining
    else:
        remaining_cost = close_amount
    pid = generate_seq_id('principals')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(
        "INSERT INTO principals (id, user_id, account_id, amount, source_type, note, created_at, record_date, investment_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (pid, user_id, account_id, abs(close_amount), 'investment_income', '投资回收：' + inv_name + ' 全额清仓', now, record_date, iid)
    )
    # 更新账户投资金额
    db.execute(
        "UPDATE accounts SET investment_amount = COALESCE(investment_amount, 0) - ? WHERE id = ? AND user_id = ?",
        (remaining_cost, account_id, user_id)
    )
    db.commit()
    record_snapshot(user_id)
    return jsonify({'success': True, 'message': '清仓成功'})

# ========== 标的库 API ==========
@app.route('/api/securities', methods=['GET'])
def get_securities():
    user_id = request.args.get('user_id', '')
    security_type = request.args.get('security_type', '').strip()
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '').strip()
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    db = get_db()
    query = "SELECT * FROM securities WHERE user_id=?"
    params = [user_id]
    if security_type:
        query += " AND security_type=?"
        params.append(security_type)
    if search:
        query += " AND (code LIKE ? OR name LIKE ?)"
        params.extend([f'%{search}%', f'%{search}%'])
    query += " ORDER BY created_at DESC"
    rows = db.execute(query, params).fetchall()
    return jsonify({'success': True, 'securities': rows_to_list(rows)})

@app.route('/api/securities/<sid>', methods=['GET'])
def get_security(sid):
    db = get_db()
    s = db.execute("SELECT * FROM securities WHERE id=?", (sid,)).fetchone()
    if not s:
        return jsonify({'success': False, 'message': '标的不存在'}), 404
    return jsonify({'success': True, 'security': row_to_dict(s)})

@app.route('/api/securities', methods=['POST'])
def add_security():
    data = request.get_json()
    user_id = data.get('user_id', '')
    code = data.get('code', '').strip()
    name = data.get('name', '').strip()
    security_type = data.get('security_type', '').strip()
    exchange = data.get('exchange', '').strip()
    description = data.get('description', '').strip()
    decimals = int(data.get('decimals', 2))
    rate = fmt(data.get('rate', 0))
    if not user_id or not name:
        return jsonify({'success': False, 'message': '缺少必填字段'}), 400
    sid = generate_seq_id('securities')
    now = datetime.now().isoformat()
    db = get_db()
    db.execute(
        "INSERT INTO securities (id, user_id, code, name, security_type, exchange, description, decimals, rate, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (sid, user_id, code, name, security_type, exchange, description, decimals, rate, now)
    )
    db.commit()
    return jsonify({'success': True, 'message': '标的添加成功',
                    'security': {'id': sid, 'user_id': user_id, 'code': code, 'name': name,
                                 'security_type': security_type, 'exchange': exchange,
                                 'description': description, 'decimals': decimals, 'rate': rate, 'created_at': now}})

@app.route('/api/securities/<sid>', methods=['PUT'])
def update_security(sid):
    data = request.get_json()
    db = get_db()
    s = db.execute("SELECT * FROM securities WHERE id=?", (sid,)).fetchone()
    if not s:
        return jsonify({'success': False, 'message': '标的不存在'}), 404
    updates = []
    params = []
    fields = ['code', 'name', 'security_type', 'exchange', 'description', 'decimals', 'rate']
    for f in fields:
        if f in data:
            updates.append(f"{f}=?")
            if f == 'decimals':
                params.append(int(data[f]))
            elif f == 'rate':
                params.append(fmt(data[f]))
            else:
                params.append(data[f].strip() if isinstance(data[f], str) else data[f])
    if updates:
        params.append(sid)
        db.execute(f"UPDATE securities SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
    s = db.execute("SELECT * FROM securities WHERE id=?", (sid,)).fetchone()
    return jsonify({'success': True, 'message': '更新成功', 'security': row_to_dict(s)})

@app.route('/api/securities/<sid>', methods=['DELETE'])
def delete_security(sid):
    db = get_db()
    s = db.execute("SELECT * FROM securities WHERE id=?", (sid,)).fetchone()
    if not s:
        return jsonify({'success': False, 'message': '标的不存在'}), 404
    # 检查是否有 investment 引用该标的
    ref_count = db.execute("SELECT COUNT(*) FROM investments WHERE security_id=?", (sid,)).fetchone()[0]
    if ref_count > 0:
        return jsonify({'success': False, 'message': f'该标的存在 {ref_count} 条投资引用，无法删除'}), 400
    db.execute("DELETE FROM securities WHERE id=?", (sid,))
    db.commit()
    return jsonify({'success': True, 'message': '删除成功'})

# ========== 公式 API ==========
@app.route('/api/formulas', methods=['GET'])
def get_formulas():
    user_id = request.args.get('user_id', '')
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    db = get_db()
    # 返回用户自定义公式 + 系统内置公式
    formulas = db.execute(
        "SELECT * FROM formulas WHERE user_id=? OR user_id='__system__' ORDER BY sort_order ASC, created_at ASC",
        (user_id,)
    ).fetchall()
    result = rows_to_list(formulas)
    # 标记内置公式
    for f in result:
        f['is_builtin'] = f.get('user_id') == '__system__'
    return jsonify({'success': True, 'formulas': result})

@app.route('/api/formulas', methods=['POST'])
def add_formula():
    data = request.get_json()
    user_id = data.get('user_id', '')
    name = data.get('name', '').strip()
    formula = data.get('formula', '').strip() or data.get('expression', '').strip()
    target_field = data.get('target_field', '').strip()
    description = data.get('description', '').strip() or data.get('note', '').strip()
    is_active = 1 if data.get('is_active', True) else 0
    sort_order = int(data.get('sort_order', 0))
    if not user_id or not name or not formula:
        return jsonify({'success': False, 'message': '缺少必填字段'}), 400
    fid = generate_seq_id('formulas')
    now = datetime.now().isoformat()
    db = get_db()
    db.execute(
        "INSERT INTO formulas (id, user_id, name, formula, target_field, description, is_active, sort_order, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (fid, user_id, name, formula, target_field, description, is_active, sort_order, now)
    )
    db.commit()
    return jsonify({'success': True, 'message': '公式添加成功',
                    'formula': {'id': fid, 'user_id': user_id, 'name': name, 'formula': formula,
                                'target_field': target_field, 'description': description,
                                'is_active': is_active, 'sort_order': sort_order, 'created_at': now,
                                'is_builtin': False}})

@app.route('/api/formulas/<fid>', methods=['PUT'])
def update_formula(fid):
    data = request.get_json()
    db = get_db()
    f = db.execute("SELECT * FROM formulas WHERE id=?", (fid,)).fetchone()
    if not f:
        return jsonify({'success': False, 'message': '公式不存在'}), 404
    # 内置公式不可修改
    if f['user_id'] == '__system__':
        return jsonify({'success': False, 'message': '内置公式不可修改'}), 403
    # 兼容前端字段名 expression->formula, note->description
    field_alias = {'expression': 'formula', 'note': 'description'}
    updates = []
    params = []
    fields = ['name', 'formula', 'target_field', 'description', 'is_active', 'sort_order']
    for field in fields:
        val = data.get(field)
        if val is None and field in field_alias:
            val = data.get(field_alias[field])
        if val is not None:
            updates.append(f"{field}=?")
            if field == 'is_active':
                params.append(1 if val else 0)
            elif field == 'sort_order':
                params.append(int(val))
            else:
                params.append(val)
    if updates:
        params.append(fid)
        db.execute(f"UPDATE formulas SET {', '.join(updates)} WHERE id=?", params)
        db.commit()
    f = db.execute("SELECT * FROM formulas WHERE id=?", (fid,)).fetchone()
    result = row_to_dict(f)
    result['is_builtin'] = False
    return jsonify({'success': True, 'message': '更新成功', 'formula': result})

@app.route('/api/formulas/<fid>', methods=['DELETE'])
def delete_formula(fid):
    db = get_db()
    f = db.execute("SELECT * FROM formulas WHERE id=?", (fid,)).fetchone()
    if not f:
        return jsonify({'success': False, 'message': '公式不存在'}), 404
    # 内置公式不可删除
    if f['user_id'] == '__system__':
        return jsonify({'success': False, 'message': '内置公式不可删除'}), 403
    db.execute("DELETE FROM formulas WHERE id=?", (fid,))
    db.commit()
    return jsonify({'success': True, 'message': '删除成功'})

# ========== 公式变量说明 API ==========
@app.route('/api/formula-vars', methods=['GET'])
def get_formula_vars():
    variables = [
        {"key": "total_principal", "label": "总本金", "description": "所有非负债账户本金余额之和"},
        {"key": "total_return", "label": "累计收益", "description": "所有收益流水的净额合计"},
        {"key": "total_assets", "label": "总资产", "description": "总本金 + 累计收益"},
        {"key": "holding_principal", "label": "持有本金", "description": "基金/股票账户的本金余额合计"},
        {"key": "total_debt", "label": "负债总额", "description": "所有负债类账户余额之和"},
        {"key": "debt_ratio", "label": "负债比率", "description": "负债总额 / 总资产 × 100%"},
        {"key": "account_count", "label": "账户数量", "description": "当前用户所有账户总数"}
    ]
    return jsonify({'success': True, 'variables': variables})

# ========== Ledger API ==========
@app.route('/api/ledger/account-flows', methods=['GET'])
def ledger_account_flows():
    """账户台账：仅保留本金相关流水，按账户筛选"""
    user_id = request.args.get('user_id', '')
    account_id = request.args.get('account_id', '').strip()
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400

    db = get_db()
    acc_filter = "AND p.account_id=?" if account_id else ""
    acc_params = (account_id,) if account_id else ()

    # 本金流水（不含投资支出/回收，避免与资产台账重复）
    p_rows = db.execute(f"""
        SELECT p.id, p.account_id, a.name as account_name,
               p.amount, p.source_type, p.note, p.record_date, p.created_at,
               'principal' as flow_type, '' as investment_name, '' as investment_type, '' as investment_direction
        FROM principals p
        LEFT JOIN accounts a ON p.account_id = a.id
        WHERE p.user_id = ? {acc_filter}
        ORDER BY p.record_date DESC, p.created_at DESC
    """, (user_id,) + acc_params).fetchall()

    result = []
    for r in p_rows:
        result.append({
            'id': r['id'], 'account_name': r['account_name'] or '-',
            'flow_type': 'principal', 'investment_name': '',
            'investment_type': '', 'investment_direction': '',
            'amount': float(r['amount']), 'source_type': r['source_type'] or '',
            'note': r['note'] or '', 'record_date': r['record_date'] or ''
        })
    # 按日期降序排列
    result.sort(key=lambda x: x['record_date'], reverse=True)
    return jsonify({'success': True, 'data': result})


@app.route('/api/ledger/asset-flows', methods=['GET'])
def ledger_asset_flows():
    """资产台账：投资流水 + 收益流水，支持按账户和投资标的筛选"""
    user_id = request.args.get('user_id', '')
    account_id = request.args.get('account_id', '').strip()
    investment_id = request.args.get('investment_id', '').strip()
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400

    db = get_db()

    # 投资流水
    inv_sql = "SELECT i.id, i.account_id, a.name as account_name, " \
              "i.name as investment_name, i.investment_type, i.direction, " \
              "i.amount, '' as source_type, i.name as note, " \
              "i.start_date as record_date, 'investment' as flow_type " \
              "FROM investments i LEFT JOIN accounts a ON i.account_id = a.id " \
              "WHERE i.user_id = ? "
    inv_params = [user_id]
    if account_id:
        inv_sql += "AND i.account_id = ? "
        inv_params.append(account_id)
    if investment_id:
        inv_sql += "AND i.id = ? "
        inv_params.append(investment_id)

    i_rows = db.execute(inv_sql + "ORDER BY i.start_date DESC, i.created_at DESC", inv_params).fetchall()

    # 收益流水
    ret_sql = "SELECT r.id, r.account_id, a.name as account_name, " \
              "i.name as investment_name, i.investment_type, '' as direction, " \
              "r.amount, r.return_type as source_type, r.note, " \
              "r.record_date, 'return' as flow_type " \
              "FROM returns r " \
              "LEFT JOIN accounts a ON r.account_id = a.id " \
              "LEFT JOIN investments i ON r.investment_id = i.id " \
              "WHERE r.user_id = ? "
    ret_params = [user_id]
    if account_id:
        ret_sql += "AND r.account_id = ? "
        ret_params.append(account_id)
    if investment_id:
        ret_sql += "AND r.investment_id = ? "
        ret_params.append(investment_id)

    r_rows = db.execute(ret_sql + "ORDER BY r.record_date DESC, r.created_at DESC", ret_params).fetchall()

    # 投资回收流水（卖出/清仓的本金回收）
    rec_sql = "SELECT p.id, p.account_id, a.name as account_name, " \
              "i.name as investment_name, i.investment_type, '' as direction, " \
              "p.amount, p.source_type, p.note, " \
              "p.record_date, 'recovery' as flow_type " \
              "FROM principals p " \
              "LEFT JOIN accounts a ON p.account_id = a.id " \
              "LEFT JOIN investments i ON p.investment_id = i.id " \
              "WHERE p.user_id = ? AND p.investment_id != '' AND p.source_type = 'investment_income' "
    rec_params = [user_id]
    if account_id:
        rec_sql += "AND p.account_id = ? "
        rec_params.append(account_id)
    if investment_id:
        rec_sql += "AND p.investment_id = ? "
        rec_params.append(investment_id)
    rec_rows = db.execute(rec_sql + "ORDER BY p.record_date DESC, p.created_at DESC", rec_params).fetchall()

    result = []
    for r in i_rows:
        result.append({
            'id': r['id'], 'account_name': r['account_name'] or '-',
            'investment_name': r['investment_name'] or '',
            'investment_type': r['investment_type'] or '',
            'investment_direction': r['direction'] or '',
            'flow_type': 'investment', 'amount': float(r['amount']),
            'source_type': r['source_type'] or '',
            'note': r['note'] or '', 'record_date': r['record_date'] or ''
        })
    for r in r_rows:
        result.append({
            'id': r['id'], 'account_name': r['account_name'] or '-',
            'investment_name': r['investment_name'] or '',
            'investment_type': r['investment_type'] or '',
            'investment_direction': '',
            'flow_type': 'return', 'amount': float(r['amount']),
            'source_type': r['source_type'] or '',
            'note': r['note'] or '', 'record_date': r['record_date'] or ''
        })
    for r in rec_rows:
        result.append({
            'id': r['id'], 'account_name': r['account_name'] or '-',
            'investment_name': r['investment_name'] or '',
            'investment_type': r['investment_type'] or '',
            'investment_direction': '',
            'flow_type': 'recovery', 'amount': -abs(float(r['amount'])),
            'source_type': r['source_type'] or '',
            'note': r['note'] or '', 'record_date': r['record_date'] or ''
        })
    result.sort(key=lambda x: x['record_date'], reverse=True)
    return jsonify({'success': True, 'data': result})

# ========== 导入导出 API ==========
@app.route('/api/export', methods=['POST'])
def export_data():
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({'success': False, 'message': '缺少 openpyxl 依赖，请安装后重试'}), 500

    data = request.get_json() or {}
    table_name = data.get('table_name', '').strip()
    export_rows = data.get('data')
    user_id = data.get('user_id', '').strip()
    is_template = data.get('template', False)

    if not table_name:
        return jsonify({'success': False, 'message': '缺少 table_name'}), 400

    # 中文表头与关键字段映射
    export_field_maps = {
        'accounts': {
            '账户名称': 'name', '账户类型': 'type', '本金余额': 'amount',
            '投资金额': 'investment_amount', '累计收益': 'total_return',
            '开户日期': 'open_date', '备注': 'remark'
        },
        'investments': {
            '账户ID': 'account_id', '名称': 'name', '投资类型': 'investment_type',
            '方向': 'direction', '投资金额': 'amount', '开始日期': 'start_date',
            '结束日期': 'end_date', '状态': 'status', '标的ID': 'security_id',
            '成本价': 'cost', '数量': 'quantity', '剩余数量': 'remaining_quantity'
        },
        'principals': {
            '账户ID': 'account_id', '金额': 'amount', '类型': 'source_type',
            '备注': 'note', '日期': 'record_date', '关联投资ID': 'investment_id'
        },
        'returns': {
            '账户ID': 'account_id', '金额': 'amount', '收益类型': 'return_type',
            '备注': 'note', '日期': 'record_date', '关联投资ID': 'investment_id'
        },
        'transfers': {
            '转出账户ID': 'from_account_id', '转入账户ID': 'to_account_id',
            '金额': 'amount', '手续费': 'fee', '划转类型': 'transfer_type', '备注': 'note'
        },
        'securities': {
            '代码': 'code', '名称': 'name', '标的类型': 'security_type',
            '交易所': 'exchange', '描述': 'description', '小数位': 'decimals', '费率': 'rate'
        },
    }

    # 模板模式：直接生成带中文表头的空模板
    if is_template:
        field_map = export_field_maps.get(table_name, {})
        if not field_map:
            return jsonify({'success': False, 'message': '该表不支持模板下载'}), 400
        wb = Workbook()
        ws = wb.active
        ws.title = table_name + '_template'
        ws.append(list(field_map.keys()))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        base64_data = base64.b64encode(output.read()).decode('utf-8')
        return jsonify({
            'success': True,
            'data': base64_data,
            'filename': f'{table_name}_template.xlsx'
        })

    db = get_db()
    # 若未传 data，则从数据库查询
    if export_rows is None:
        if not user_id:
            return jsonify({'success': False, 'message': '缺少 user_id 或 data'}), 400
        allowed_tables = {
            'accounts': "SELECT * FROM accounts WHERE user_id=?",
            'principals': "SELECT * FROM principals WHERE user_id=?",
            'returns': "SELECT * FROM returns WHERE user_id=?",
            'transfers': "SELECT * FROM transfers WHERE user_id=?",
            'snapshots': "SELECT * FROM snapshots WHERE user_id=?",
            'investments': "SELECT * FROM investments WHERE user_id=?",
            'formulas': "SELECT * FROM formulas WHERE user_id=?",
            'dict_items': "SELECT * FROM dict_items WHERE user_id=? OR user_id='__system__'",
            'securities': "SELECT * FROM securities WHERE user_id=?",
        }
        if table_name not in allowed_tables:
            return jsonify({'success': False, 'message': '不支持的表名'}), 400
        sql = allowed_tables[table_name]
        params = (user_id,) if '__system__' not in sql else ()
        rows = db.execute(sql, params).fetchall()
        export_rows = rows_to_list(rows)

    if not export_rows:
        return jsonify({'success': False, 'message': '没有数据可导出'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = table_name

    field_map = export_field_maps.get(table_name, {})
    if field_map:
        # 使用中文表头，只导出关键字段
        headers = list(field_map.keys())
        en_headers = list(field_map.values())
        ws.append(headers)
        for row in export_rows:
            ws.append([row.get(h, '') for h in en_headers])
    else:
        # 无映射时全量导出
        headers = list(export_rows[0].keys())
        ws.append(headers)
        for row in export_rows:
            ws.append([row.get(h, '') for h in headers])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    base64_data = base64.b64encode(output.read()).decode('utf-8')

    return jsonify({
        'success': True,
        'data': base64_data,
        'filename': f'{table_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    })

@app.route('/api/import', methods=['POST'])
def import_data():
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({'success': False, 'message': '缺少 openpyxl 依赖，请安装后重试'}), 500

    data = request.get_json() or {}
    table_name = data.get('table_name', '').strip()
    xlsx_base64 = data.get('data', '').strip()
    user_id = data.get('user_id', '').strip()

    if not table_name:
        return jsonify({'success': False, 'message': '缺少 table_name'}), 400
    if not xlsx_base64:
        return jsonify({'success': False, 'message': '缺少 data (xlsx base64)'}), 400
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400

    # 中文表头 -> 英文字段映射（复用 export 的映射）
    import_field_maps = {
        'accounts': {
            '账户名称': 'name', '账户类型': 'type', '本金余额': 'amount',
            '投资金额': 'investment_amount', '累计收益': 'total_return',
            '开户日期': 'open_date', '备注': 'remark'
        },
        'investments': {
            '账户ID': 'account_id', '名称': 'name', '投资类型': 'investment_type',
            '方向': 'direction', '投资金额': 'amount', '开始日期': 'start_date',
            '结束日期': 'end_date', '状态': 'status', '标的ID': 'security_id',
            '成本价': 'cost', '数量': 'quantity', '剩余数量': 'remaining_quantity'
        },
        'principals': {
            '账户ID': 'account_id', '金额': 'amount', '类型': 'source_type',
            '备注': 'note', '日期': 'record_date', '关联投资ID': 'investment_id'
        },
        'returns': {
            '账户ID': 'account_id', '金额': 'amount', '收益类型': 'return_type',
            '备注': 'note', '日期': 'record_date', '关联投资ID': 'investment_id'
        },
        'transfers': {
            '转出账户ID': 'from_account_id', '转入账户ID': 'to_account_id',
            '金额': 'amount', '手续费': 'fee', '划转类型': 'transfer_type', '备注': 'note'
        },
        'securities': {
            '代码': 'code', '名称': 'name', '标的类型': 'security_type',
            '交易所': 'exchange', '描述': 'description', '小数位': 'decimals', '费率': 'rate'
        },
    }

    allowed_tables = {
        'accounts': ['id', 'user_id', 'name', 'type', 'amount', 'investment_amount', 'total_return', 'open_date', 'remark', 'created_at'],
        'principals': ['id', 'user_id', 'account_id', 'amount', 'source_type', 'note', 'created_at', 'record_date', 'investment_id', 'transfer_id'],
        'returns': ['id', 'user_id', 'account_id', 'amount', 'return_type', 'note', 'created_at', 'record_date', 'investment_id'],
        'transfers': ['id', 'user_id', 'from_account_id', 'to_account_id', 'amount', 'fee', 'transfer_type', 'note', 'created_at'],
        'investments': ['id', 'user_id', 'account_id', 'name', 'investment_type', 'direction', 'amount', 'start_date', 'end_date', 'status', 'security_id', 'cost', 'quantity', 'remaining_quantity', 'created_at'],
        'formulas': ['id', 'user_id', 'name', 'formula', 'target_field', 'description', 'is_active', 'sort_order', 'created_at'],
        'dict_items': ['id', 'user_id', 'dict_type', 'label', 'value', 'sort_order', 'created_at'],
        'securities': ['id', 'user_id', 'code', 'name', 'security_type', 'exchange', 'description', 'decimals', 'rate', 'created_at'],
    }

    if table_name not in allowed_tables:
        return jsonify({'success': False, 'message': '不支持的表名'}), 400

    try:
        xlsx_bytes = base64.b64decode(xlsx_base64)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = list(next(rows_iter))

        # 中文表头映射
        field_map = import_field_maps.get(table_name, {})
        headers = []
        for h in raw_headers:
            h_str = str(h).strip() if h is not None else ''
            if h_str in field_map:
                headers.append(field_map[h_str])
            elif h_str in allowed_tables[table_name]:
                headers.append(h_str)
            else:
                headers.append(h_str)

        imported = 0
        skipped = 0
        db = get_db()
        for row in rows_iter:
            row_dict = dict(zip(headers, row))
            # 强制绑定 user_id
            row_dict['user_id'] = user_id
            # 若 id 为空则生成递增ID
            if not row_dict.get('id'):
                row_dict['id'] = generate_seq_id(table_name)
            # 过滤出有效字段
            valid_cols = [c for c in allowed_tables[table_name] if c in row_dict]
            if not valid_cols:
                skipped += 1
                continue
            cols_str = ', '.join(valid_cols)
            placeholders = ', '.join(['?' for _ in valid_cols])
            values = [row_dict.get(c, '') for c in valid_cols]
            try:
                db.execute(f"INSERT OR REPLACE INTO {table_name} ({cols_str}) VALUES ({placeholders})", values)
                imported += 1
            except Exception:
                skipped += 1
        db.commit()
        return jsonify({'success': True, 'message': f'导入完成：成功 {imported} 条，跳过 {skipped} 条', 'imported': imported, 'skipped': skipped})
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败：{str(e)}'}), 500

# ========== 重置数据 API ==========
@app.route('/api/reset-data', methods=['POST'])
def reset_data():
    data = request.get_json() or {}
    user_id = data.get('user_id', '').strip()
    if not user_id:
        return jsonify({'success': False, 'message': '缺少 user_id'}), 400
    db = get_db()
    user = db.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    print(f"[RESET] 用户 {user_id} 请求重置数据，已确认执行")
    db.execute("DELETE FROM transfers WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM returns WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM principals WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM investments WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM formulas WHERE user_id=? AND user_id != '__system__'", (user_id,))
    db.execute("DELETE FROM securities WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM accounts WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM snapshots WHERE user_id=?", (user_id,))
    db.execute("DELETE FROM dict_items WHERE user_id=? AND user_id != '__system__'", (user_id,))
    db.commit()
    return jsonify({'success': True, 'message': '数据重置成功，所有账户、本金、收益、划转、投资、公式、标的、快照数据已清除'})

# ========== 单账户已实现盈亏明细 API ==========
@app.route('/api/accounts/<acc_id>/realized-returns', methods=['GET'])
def get_account_realized_returns(acc_id):
    """返回账户每笔投资的累计卖出/清仓盈亏明细"""
    user_id = request.args.get('user_id', '')
    if not user_id or not acc_id:
        return jsonify({'success': False, 'message': '缺少 user_id 或 acc_id'}), 400
    db = get_db()
    # 先查询每笔投资的基础信息
    inv_rows = db.execute("""
        SELECT i.id as investment_id, i.name as investment_name, i.investment_type,
               i.amount as inv_amount, i.quantity as inv_qty, i.remaining_quantity
        FROM investments i
        WHERE i.account_id = ? AND i.user_id = ?
    """, (acc_id, user_id)).fetchall()

    result = []
    for inv in inv_rows:
        iid = inv['investment_id']
        inv_amount = float(inv['inv_amount'] or 0)
        inv_qty = float(inv['inv_qty'] or 0)
        remaining = float(inv['remaining_quantity'] or 0)
        sold_qty = inv_qty - remaining
        if sold_qty <= 0:
            continue
        # 查询该投资的总卖出金额
        total_sell = db.execute(
            "SELECT COALESCE(SUM(amount), 0) FROM principals WHERE investment_id=? AND source_type='investment_income'",
            (iid,)
        ).fetchone()[0]
        total_sell = float(total_sell or 0)
        if total_sell <= 0:
            continue
        # 计算成本（平均成本法）
        if inv_qty > 0:
            sold_cost = inv_amount / inv_qty * sold_qty
        else:
            sold_cost = 0
        realized = total_sell - sold_cost
        result.append({
            'investment_id': iid,
            'investment_name': inv['investment_name'] or '',
            'investment_type': inv['investment_type'] or '',
            'sell_amount': fmt(total_sell),
            'sold_cost': fmt(sold_cost),
            'realized_return': fmt(realized),
            'formula': f'累计卖出金额({fmt(total_sell)}) - 成本({fmt(sold_cost)}) = 已实现盈亏({fmt(realized)})'
        })
    return jsonify({'success': True, 'data': result})

# ========== 单账户资产历史 API ==========
@app.route('/api/accounts/<acc_id>/history', methods=['GET'])
def get_account_history(acc_id):
    user_id = request.args.get('user_id', '')
    if not user_id or not acc_id:
        return jsonify({'success': False, 'message': '缺少 user_id 或 acc_id'}), 400
    db = get_db()
    acc = db.execute("SELECT * FROM accounts WHERE id=? AND user_id=?", (acc_id, user_id)).fetchone()
    if not acc:
        return jsonify({'success': False, 'message': '账户不存在'}), 404

    principals = db.execute(
        "SELECT record_date, amount, source_type FROM principals WHERE account_id=? AND user_id=? ORDER BY record_date ASC, created_at ASC",
        (acc_id, user_id)
    ).fetchall()

    returns = db.execute(
        "SELECT record_date, amount FROM returns WHERE account_id=? AND user_id=? ORDER BY record_date ASC, created_at ASC",
        (acc_id, user_id)
    ).fetchall()

    today = datetime.now().strftime('%Y-%m-%d')
    all_dates = set()
    for p in principals:
        if p['record_date']:
            all_dates.add(p['record_date'])
    for r in returns:
        if r['record_date']:
            all_dates.add(r['record_date'])
    all_dates.add(today)

    if not all_dates:
        return jsonify({'success': True, 'dates': [], 'principal_balance': [], 'return_cumulative': [], 'total': []})

    min_date = min(all_dates)
    max_date = min(today, max(all_dates))

    start_dt = datetime.strptime(min_date, '%Y-%m-%d')
    end_dt = datetime.strptime(max_date, '%Y-%m-%d')
    date_list = []
    current = start_dt
    while current <= end_dt:
        date_list.append(current.strftime('%Y-%m-%d'))
        current += timedelta(days=1)
    if len(date_list) > 30:
        date_list = date_list[-30:]

    principal_by_date = {}
    for p in principals:
        d = p['record_date'] or ''
        if not d:
            continue
        principal_by_date[d] = principal_by_date.get(d, 0) + p['amount']

    return_by_date = {}
    for r in returns:
        d = r['record_date'] or ''
        if not d:
            continue
        return_by_date[d] = return_by_date.get(d, 0) + r['amount']

    non_initial_change = 0.0
    for p in principals:
        if p['source_type'] != 'initial':
            non_initial_change += p['amount']
    actual_initial = acc['amount'] - non_initial_change

    principal_balance_list = []
    return_cumulative_list = []
    total_list = []
    cum_principal = actual_initial
    cum_return = 0.0
    for d in date_list:
        day_change = 0.0
        for p in principals:
            if p['record_date'] == d and p['source_type'] != 'initial':
                day_change += p['amount']
        cum_principal += day_change
        cum_return += return_by_date.get(d, 0)
        principal_balance_list.append(fmt(cum_principal))
        return_cumulative_list.append(fmt(cum_return))
        total_list.append(fmt(cum_principal + cum_return))

    return jsonify({
        'success': True,
        'dates': date_list,
        'principal_balance': principal_balance_list,
        'return_cumulative': return_cumulative_list,
        'total': total_list
    })

# ========== 静态文件 ==========
@app.route('/')
def serve_index():
    return send_from_directory('static', 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)

# 模块级初始化：确保 PythonAnywhere WSGI 导入时也执行
init_db()

if __name__ == '__main__':
    print('=' * 50)
    print('  个人多账户资产管理系统 v5.0 (SQLite)')
    print('  访问 http://localhost:5000')
    print('=' * 50)
    app.run(host='0.0.0.0', port=5000, debug=True)
